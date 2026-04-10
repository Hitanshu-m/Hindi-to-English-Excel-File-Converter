from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
import re
import json
import os
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate

app = Flask(__name__)

DICT_FILE = "dictionary.json"

# Load dictionary
if os.path.exists(DICT_FILE):
    with open(DICT_FILE, "r", encoding="utf-8") as f:
        dictionary = json.load(f)
else:
    dictionary = {}

def is_devanagari(text):
    return isinstance(text, str) and re.search(r'[\u0900-\u097F]', text)

def save_dictionary():
    with open(DICT_FILE, "w", encoding="utf-8") as f:
        json.dump(dictionary, f, ensure_ascii=False, indent=2)

def convert_word(word):
    clean = re.sub(r'[^\u0900-\u097F]', '', word)

    if clean in dictionary:
        return dictionary[clean]

    if is_devanagari(word):
        try:
            converted = transliterate(word, sanscript.DEVANAGARI, sanscript.HK).title()
            dictionary[clean] = converted
            return converted
        except:
            return word

    return word

def improve_sentence(words):
    result = []
    i = 0
    while i < len(words):
        if i < len(words)-1 and words[i] == "of" and words[i+1] == "house":
            result.append("house")
            result.append("of")
            i += 2
            continue

        result.append(words[i])
        i += 1

    return " ".join(result)

def convert_text(text):
    if not isinstance(text, str):
        return text

    lines = text.split('\n')
    new_lines = []

    for line in lines:
        words = line.split()
        converted_words = [convert_word(w) for w in words]
        sentence = improve_sentence(converted_words)
        new_lines.append(sentence)

    return "\n".join(new_lines)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/upload", methods=["POST"])
def upload():
    file = request.files["file"]

    if not file:
        return "No file uploaded"

    input_path = "input.xlsx"
    output_path = "converted.xlsx"

    file.save(input_path)

    wb = load_workbook(input_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.value = convert_text(cell.value)

    wb.save(output_path)
    save_dictionary()

    return send_file(output_path, as_attachment=True, download_name="converted.xlsx")

# Render compatible
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)